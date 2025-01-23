
from openpyxl import load_workbook

DEFAULT_FILE = "Applications Applied.xlsx"
TITLE_ROW = 1   #change if your titles are not in this row
STARTING_COLUMN = 2 #change if your format is different
STARTING_ROW = 2    #change if your format is different

#adds a new row to the file, prompting the user for data for each column
def addRow(file):


    #load excel file
    workbook = load_workbook(file)

    #open workbook
    sheet = workbook.active

    currColumn = STARTING_COLUMN

    currRow = STARTING_ROW
    currCell = sheet.cell(currRow, currColumn)
    #find the lowest empty row, since the maximum row may not be correct
    while(currCell.value):
        currRow+=1
        currCell = sheet.cell(currRow, currColumn)
    

    updateRow = sheet.max_row+1
    if(updateRow > currRow):
        updateRow = currRow
        
    while(currColumn<sheet.max_column+1):
        print("Enter the", sheet.cell(row= TITLE_ROW, column= currColumn).value)
        currInput = input()
        currCell = sheet.cell(updateRow, currColumn)
        currCell.value = currInput

        currColumn+=1

    #save file
    workbook.save(file)
    workbook.close()

#updates the desired row in the file, prompting user for new data
def updateRow(file):
    
    #load excel file
    workbook = load_workbook(file)
    #open workbook
    sheet = workbook.active

    #prompt for position or company name
    print("\nNote: Enter specific information to find fewer applications:")
    rows = validCompPos(file)

    if(rows):
        print("")
        for i in range(len(rows)):
            currColumn = STARTING_COLUMN
            while(currColumn<sheet.max_column+1):
                print(currColumn-1, sheet.cell(row= TITLE_ROW, column= currColumn).value + ":", sheet.cell(row= rows[i], column = currColumn).value)
                currColumn+=1
            print("")

            print("Update this information? \nType \"Yes\" to Update \nType \"Back\" to return to previous Application \nType \"Menu\" to return to main menu \nAny other input or noninput will move you to the next application if there is one")
            inputs = input()

            if(inputs.lower() == "menu"):
                workbook.close()
                return
            elif(inputs.lower() == "back"):
                i -= 2
            elif(inputs.lower() == "yes" or inputs.lower() == "y"):
                updated = False
                while(not updated):
                    print("Which row to update?")
                    inputs = input()

                    upCol = getCol(file, inputs)
                    if(inputs.isnumeric() and int(inputs)+1 <= sheet.max_column+1 and int(inputs)+1 >= STARTING_COLUMN):
                        upCol = int(inputs)+1
                        print("What should it be updated to?")
                        inputs = input()

                        sheet.cell(rows[i], upCol).value = inputs
                        updated = True
                        workbook.save(file)
                    elif(upCol):
                        print("What should it be updated to?")
                        inputs = input()

                        sheet.cell(rows[i], upCol).value = inputs
                        updated = True
                        workbook.save(file)


    #save file
    workbook.save(file)
    workbook.close()
    
#gets the current row of the input value to check
def getCol(file, check):
    #load excel file
    workbook = load_workbook(file)
    #open workbook
    sheet = workbook.active

    currCol = STARTING_COLUMN
    currTitle = sheet.cell(TITLE_ROW, currCol)
    while(currTitle.value):
        if(currTitle.value.lower() == check.lower()):
            workbook.close()
            return currCol

        currCol+=1
        currTitle = sheet.cell(TITLE_ROW, currCol)
        
        
    return False

def viewRow(file):
    
    ###Add a flag to all for partial matching rather than full matching? ###

    #load excel file
    workbook = load_workbook(file)
    #open workbook
    sheet = workbook.active

    posRow = validCompPos(file)
        
    
    for values in posRow:
        currColumn = STARTING_COLUMN
        while(currColumn<sheet.max_column+1):
            print(sheet.cell(row= TITLE_ROW, column= currColumn).value + ":", sheet.cell(row= values, column = currColumn).value)
            currColumn+=1
        print("")

    print("Number of applications found:", len(posRow))
    #save file
    workbook.save(file)
    workbook.close()
    
#Determines if the input company or position name is valid or exits back to main menu otherwise
#If it is valid, gets the row or rows that the company or position name resides in the excel file
def validCompPos(file):
    inputs = ""
    print("Which Company or Position would you like to view?")
    inputs = input()

    if(inputs.lower() == "back"):
        return
    posRow = getRow(file, inputs)
    while(not posRow):
        print("Row is not in file! Please double check your spelling and try again or type 'back' to return to menu.")
        inputs = input()


        if(inputs.lower() == "back"):
            return
        else:
            posRow = getRow(file, inputs)

    return posRow
    
#returns the row or rows of the desired position
def getRow(file, position):

    #load excel file
    workbook = load_workbook(file)
    #open workbook
    sheet = workbook.active

    rows = []

    companyCol = STARTING_COLUMN
    positionCol = STARTING_COLUMN+1
    currRow = STARTING_ROW

    currCompanyCell = sheet.cell(currRow, companyCol)
    currPositionCell = sheet.cell(currRow, positionCol)
    while(currCompanyCell.value and currPositionCell.value):
        if(currCompanyCell.value.lower() == position.lower() or currPositionCell.value.lower() == position.lower()):
            rows.append(currRow)

        currRow+=1
        currCompanyCell = sheet.cell(currRow, companyCol)
        currPositionCell = sheet.cell(currRow, positionCol)
        
        
    workbook.close()
    if(len(rows) > 0):
        return rows
    else:
        return False

#Prints out the number of rows (applications completed) in the excel file
def numAppsComp(file):
    #load excel file
    workbook = load_workbook(file, data_only=True)
    #open workbook
    sheet = workbook.active
    #Print value in first cell, which is the number of rows written to (number of applications)
    print(sheet.max_row -1)
    workbook.close()

#main menu that runs
def main():
    default = False
    print("Welcome to the Application Adder!! \nIf adding to a file that is not", DEFAULT_FILE, "please type it now, otherwise hit enter.")
    x = input()
    filenames = ""
    if(x != ""):
        filenames = x + ".xlsx"
    else:
        filenames = DEFAULT_FILE

    
    inputs = ""
    while(inputs.upper() != "0"):
        print("NOTE!! If file is open, no additions can be added only viewing!")
        print("1: Add new row to spreadsheet? \n2: Update existing row? \n3: View Row Data? \n4: View Number of Applications \n0: Exit Program")
        inputs = input()
        if(inputs == "1"):
            addRow(filenames)
        elif(inputs == "2"):
            updateRow(filenames)
            print("Update Completed")
        elif(inputs == "3"):
            viewRow(filenames)
        elif(inputs == "4"):
            numAppsComp(filenames)

    return


if __name__ == "__main__":
    main()